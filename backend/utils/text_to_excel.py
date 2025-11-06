# utils/text_to_excel.py
import re
from typing import Dict, List, Tuple
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

def parse_any_format_to_excel(content: str, output_path: str) -> str:
    """
    Parse markdown, JSON, or plain text from LLM and convert to Excel.
    Handles any format intelligently.
    
    Args:
        content: Raw text from LLM (markdown, JSON, or plain text)
        output_path: Path to save Excel file
    
    Returns:
        Path to created Excel file
    """
    
    # Try to detect format and parse
    sections = parse_content(content)
    
    # Create Excel workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Extracted Guidelines"
    
    # Styling
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    section_fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
    section_font = Font(bold=True, size=11)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Write headers
    ws['A1'] = "Section"
    ws['B1'] = "Subsection/Rule"
    ws['C1'] = "Details"
    
    for cell in ['A1', 'B1', 'C1']:
        ws[cell].fill = header_fill
        ws[cell].font = header_font
        ws[cell].alignment = Alignment(horizontal='center', vertical='center')
        ws[cell].border = border
    
    # Write data
    row = 2
    for section_name, section_data in sections.items():
        # Section header row
        ws[f'A{row}'] = section_name
        ws[f'A{row}'].fill = section_fill
        ws[f'A{row}'].font = section_font
        ws[f'A{row}'].border = border
        ws[f'B{row}'].border = border
        ws[f'C{row}'].border = border
        
        # Section summary if exists
        if 'summary' in section_data:
            ws[f'B{row}'] = "Summary"
            ws[f'C{row}'] = section_data['summary']
            ws[f'B{row}'].fill = section_fill
            ws[f'C{row}'].fill = section_fill
            ws[f'B{row}'].font = Font(italic=True)
            ws[f'C{row}'].font = Font(italic=True)
            row += 1
        else:
            row += 1
        
        # Section items
        for key, value in section_data.items():
            if key != 'summary':
                ws[f'A{row}'] = ""  # Empty for sub-items
                ws[f'B{row}'] = key
                ws[f'C{row}'] = value
                
                ws[f'A{row}'].border = border
                ws[f'B{row}'].border = border
                ws[f'C{row}'].border = border
                
                row += 1
    
    # Adjust column widths
    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 40
    ws.column_dimensions['C'].width = 80
    
    # Enable text wrapping
    for row_cells in ws.iter_rows(min_row=1, max_row=ws.max_row):
        for cell in row_cells:
            cell.alignment = Alignment(wrap_text=True, vertical='top')
    
    # Save workbook
    wb.save(output_path)
    print(f"✅ Excel file created: {output_path}")
    
    return output_path


def parse_content(content: str) -> Dict[str, Dict[str, str]]:
    """
    Intelligently parse content in any format (markdown, text, JSON).
    
    Returns structured dict:
    {
        "Section Name": {
            "summary": "Section summary",
            "Rule 1": "Rule details",
            "Rule 2": "More details"
        }
    }
    """
    import json
    
    # Try JSON first
    try:
        data = json.loads(content.strip())
        if isinstance(data, dict):
            print("✅ Detected JSON format")
            return flatten_json(data)
    except:
        pass
    
    # Parse as markdown/text
    print("✅ Parsing as markdown/text format")
    return parse_markdown(content)


def flatten_json(data: dict) -> Dict[str, Dict[str, str]]:
    """Flatten nested JSON to section-based structure"""
    result = {}
    
    for key, value in data.items():
        if isinstance(value, dict):
            result[key] = {}
            for sub_key, sub_value in value.items():
                if isinstance(sub_value, (str, int, float)):
                    result[key][sub_key] = str(sub_value)
                else:
                    result[key][sub_key] = str(sub_value)
        elif isinstance(value, (str, int, float)):
            if "General" not in result:
                result["General"] = {}
            result["General"][key] = str(value)
    
    return result


def parse_markdown(content: str) -> Dict[str, Dict[str, str]]:
    """Parse markdown-style content"""
    sections = {}
    current_section = "General"
    current_subsection = None
    
    lines = content.split('\n')
    
    for line in lines:
        line = line.strip()
        if not line:
            continue
        
        # Detect section headers (various formats)
        # **Bold Text** or ## Heading
        if line.startswith('**') and line.endswith('**'):
            current_section = line.strip('*').strip()
            sections[current_section] = {}
            current_subsection = None
            
        elif line.startswith('##'):
            current_section = line.lstrip('#').strip()
            sections[current_section] = {}
            current_subsection = None
            
        # Detect subsections
        elif line.startswith('*') and ':' in line:
            # Format: * Subsection: Details
            parts = line.lstrip('*').strip().split(':', 1)
            if len(parts) == 2:
                subsection_name = parts[0].strip().lstrip('**').rstrip('**')
                subsection_value = parts[1].strip()
                
                if current_section not in sections:
                    sections[current_section] = {}
                
                sections[current_section][subsection_name] = subsection_value
                current_subsection = subsection_name
        
        # Continuation of previous subsection
        elif current_subsection and line.startswith(' '):
            if current_section in sections and current_subsection in sections[current_section]:
                sections[current_section][current_subsection] += " " + line.strip()
        
        # Bullet point without colon
        elif line.startswith('*'):
            bullet_text = line.lstrip('*').strip()
            if current_section not in sections:
                sections[current_section] = {}
            
            # Generate key
            bullet_key = f"Item {len(sections[current_section]) + 1}"
            sections[current_section][bullet_key] = bullet_text
    
    # If no sections found, create one general section
    if not sections:
        sections["Extracted Content"] = {"Content": content}
    
    return sections