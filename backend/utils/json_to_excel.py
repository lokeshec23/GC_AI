# backend/utils/json_to_excel.py

import json
from typing import List, Dict
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

def dynamic_json_to_excel(json_data: List[Dict], output_path: str) -> str:
    """
    Dynamically converts a list of JSON objects (dictionaries) into a
    formatted Excel file. The columns are inferred from the keys of the
    first object in the list.

    Args:
        json_data: A list of dictionaries, where each dictionary represents a row.
        output_path: The full path where the Excel file will be saved.

    Returns:
        The path to the created Excel file.
    """
    
    if not json_data:
        print("⚠️ No data provided to write to Excel. Creating an empty file.")
        wb = Workbook()
        ws = wb.active
        ws['A1'] = "No structured data was extracted or generated."
        wb.save(output_path)
        return output_path

    print(f"📊 Dynamically converting {len(json_data)} items to Excel...")

    # Infer headers from the keys of the first valid object in the list
    try:
        headers = list(json_data[0].keys())
        print(f"   Inferred Headers: {headers}")
    except (IndexError, AttributeError):
        print("⚠️ JSON data is empty or not a list of dictionaries. Creating an empty file.")
        return dynamic_json_to_excel([], output_path)


    wb = Workbook()
    ws = wb.active
    ws.title = "Extraction Results"

    # --- Styling ---
    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    border_thin = Border(
        left=Side(style='thin', color='D9D9D9'),
        right=Side(style='thin', color='D9D9D9'),
        top=Side(style='thin', color='D9D9D9'),
        bottom=Side(style='thin', color='D9D9D9')
    )

    # --- Write Headers ---
    for col_num, header_title in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        # Format the header title (e.g., "major_section" -> "Major Section")
        cell.value = str(header_title).replace("_", " ").title()
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = border_thin

    # --- Write Data Rows ---
    for row_num, item in enumerate(json_data, 2):
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=row_num, column=col_num)
            value = item.get(header)
            
            # Convert complex types (lists, dicts) to a readable string format
            if isinstance(value, (list, dict)):
                try:
                    cell.value = json.dumps(value, indent=2)
                except TypeError:
                    cell.value = str(value)
            else:
                cell.value = str(value) if value is not None else ""
            
            cell.border = border_thin
            cell.alignment = Alignment(wrap_text=True, vertical='top')

    # --- Auto-fit Column Widths ---
    for col_num, header in enumerate(headers, 1):
        column_letter = get_column_letter(col_num)
        # Set a reasonable default width, as auto-sizing can be slow and imperfect
        ws.column_dimensions[column_letter].width = 35

    # Freeze the header row so it's always visible when scrolling
    ws.freeze_panes = 'A2'
    
    # --- Save Workbook ---
    try:
        wb.save(output_path)
        print(f"✅ Dynamic Excel file created successfully: {output_path}")
    except Exception as e:
        print(f"❌ Failed to save Excel file: {e}")
        raise

    return output_path