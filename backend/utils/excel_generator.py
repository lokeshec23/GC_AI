# utils/excel_generator.py
import json
from typing import Dict, Any
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
import os

def json_to_excel(json_data: dict, output_path: str) -> str:
    """
    Convert JSON extraction results to formatted Excel file.
    
    Args:
        json_data: Dictionary containing extraction results
        output_path: Full path for output Excel file
    
    Returns:
        Path to created Excel file
    """
    wb = Workbook()
    
    # Remove default sheet
    if "Sheet" in wb.sheetnames:
        wb.remove(wb["Sheet"])
    
    # Create sheet with rules data
    ws = wb.create_sheet("Extracted Rules")
    
    # Header styling
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    
    # Write headers
    ws['A1'] = "Section"
    ws['B1'] = "Subsection/Rule"
    ws['C1'] = "Details"
    
    for cell in ['A1', 'B1', 'C1']:
        ws[cell].fill = header_fill
        ws[cell].font = header_font
        ws[cell].alignment = Alignment(horizontal='center', vertical='center')
    
    # Write data
    row = 2
    
    def write_nested_data(data: Any, section_name: str = "", indent_level: int = 0):
        """Recursively write nested JSON data to Excel"""
        nonlocal row
        
        if isinstance(data, dict):
            for key, value in data.items():
                if key == "summary":
                    # Write summary for main sections
                    ws[f'A{row}'] = section_name
                    ws[f'B{row}'] = "Summary"
                    ws[f'C{row}'] = value
                    
                    # Style summary row
                    ws[f'A{row}'].font = Font(bold=True, size=11)
                    ws[f'A{row}'].fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
                    ws[f'B{row}'].fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
                    ws[f'C{row}'].fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
                    
                    row += 1
                elif isinstance(value, str):
                    # Write subsection rules
                    ws[f'A{row}'] = section_name
                    ws[f'B{row}'] = key
                    ws[f'C{row}'] = value
                    row += 1
                elif isinstance(value, (dict, list)):
                    # Recurse 