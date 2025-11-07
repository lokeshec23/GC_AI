# compare/processor.py
import os
import json
import tempfile
from typing import List, Dict
from utils.excel_reader import read_excel_to_json, excel_to_text_summary
from utils.llm_provider import LLMProvider
from utils.json_to_excel import validate_and_convert_to_excel
from utils.progress import update_progress

def process_comparison_background(
    session_id: str,
    file1_path: str,
    file2_path: str,
    file1_name: str,
    file2_name: str,
    user_settings: dict,
    model_provider: str,
    model_name: str,
    custom_prompt: str
):
    """Background task for comparing two Excel files"""
    excel_path = None
    
    try:
        print(f"\n{'='*60}")
        print(f"🔄 Comparison started for session: {session_id[:8]}")
        print(f"📄 File 1: {file1_name}")
        print(f"📄 File 2: {file2_name}")
        print(f"🤖 Model: {model_provider}/{model_name}")
        print(f"{'='*60}\n")

        # STEP 1: Read Excel Files (0% → 30%)
        update_progress(session_id, 5, "Reading first Excel file...")
        
        data1 = read_excel_to_json(file1_path, "Guideline 1")
        update_progress(session_id, 15, f"✅ Read {len(data1)} rows from first file")
        
        update_progress(session_id, 20, "Reading second Excel file...")
        data2 = read_excel_to_json(file2_path, "Guideline 2")
        update_progress(session_id, 30, f"✅ Read {len(data2)} rows from second file")
        
        print(f"✅ File 1: {len(data1)} rows")
        print(f"✅ File 2: {len(data2)} rows\n")

        # STEP 2: Convert to Text Summaries (30% → 40%)
        update_progress(session_id, 35, "Preparing data for comparison...")
        
        text_summary1 = excel_to_text_summary(data1, "GUIDELINE 1")
        text_summary2 = excel_to_text_summary(data2, "GUIDELINE 2")
        
        combined_context = f"{text_summary1}\n\n{text_summary2}"
        
        update_progress(session_id, 40, "✅ Data prepared for LLM")
        print(f"✅ Context prepared: {len(combined_context)} characters\n")

        # STEP 3: LLM Comparison (40% → 80%)
        update_progress(session_id, 45, f"Initializing {model_provider} LLM...")
        
        # Get API key
         # Initialize LLM based on provider
        if model_provider == "openai":
            api_key = user_settings.get("openai_api_key")
            endpoint = user_settings.get("openai_endpoint")
            deployment = user_settings.get("openai_deployment")
            
            if not api_key or not endpoint or not deployment:
                raise ValueError("OpenAI credentials not configured in Settings")
            
            llm = LLMProvider(
                provider="openai",
                api_key=api_key,
                model=model_name,
                temperature=user_settings.get("temperature", 0.7),
                max_tokens=user_settings.get("max_output_tokens", 8192),
                top_p=user_settings.get("top_p", 1.0),
                stop_sequences=user_settings.get("stop_sequences", []),
                azure_endpoint=endpoint,
                azure_deployment=deployment,
            )
            
        elif model_provider == "gemini":
            api_key = user_settings.get("gemini_api_key")
            if not api_key:
                raise ValueError("Gemini API key not configured in Settings")
            
            llm = LLMProvider(
                provider="gemini",
                api_key=api_key,
                model=model_name,
                temperature=user_settings.get("temperature", 0.7),
                max_tokens=user_settings.get("max_output_tokens", 8192),
                top_p=user_settings.get("top_p", 1.0),
                stop_sequences=user_settings.get("stop_sequences", [])
            )
        else:
            raise ValueError(f"Unsupported provider: {model_provider}")
        
        update_progress(session_id, 50, "Comparing guidelines with LLM...")
        
        # Build full prompt
        full_prompt = f"""{custom_prompt}

### GUIDELINE DATA TO COMPARE

{combined_context}

### INSTRUCTIONS
Compare the two guidelines above and output a JSON array with comparison results.
Each object should have:
- "category": The comparison category (e.g., "Added in Guideline 2", "Removed from Guideline 1", "Modified", "Unchanged")
- "section": The section or rule being compared
- "guideline1_value": Value/rule from Guideline 1 (or empty if not present)
- "guideline2_value": Value/rule from Guideline 2 (or empty if not present)
- "difference": Brief description of the difference or "Same" if identical

Output ONLY the JSON array. No explanations.
"""
        
        print(f"📤 Sending comparison request to LLM...")
        response = llm.generate(full_prompt)
        
        update_progress(session_id, 75, "✅ LLM comparison complete")
        print(f"✅ Got response: {len(response)} characters\n")

        # STEP 4: Parse and Validate Response (75% → 85%)
        update_progress(session_id, 78, "Parsing comparison results...")
        
        comparison_data = parse_comparison_json(response)
        
        print(f"📊 Comparison results: {len(comparison_data)} items")
        update_progress(session_id, 85, f"✅ Parsed {len(comparison_data)} comparison items")

        # STEP 5: Convert to Excel (85% → 95%)
        update_progress(session_id, 87, "Converting to Excel...")
        
        excel_path = tempfile.NamedTemporaryFile(
            delete=False,
            suffix=".xlsx",
            prefix=f"comparison_{session_id[:8]}_"
        ).name
        
        # Convert comparison data to Excel
        create_comparison_excel(comparison_data, excel_path, file1_name, file2_name)
        
        file_size = os.path.getsize(excel_path)
        update_progress(session_id, 95, f"✅ Excel generated ({file_size:,} bytes)")

        # STEP 6: Complete (95% → 100%)
        update_progress(session_id, 100, "✅ Comparison complete!")
        
        print(f"{'='*60}")
        print(f"✅ COMPARISON COMPLETE")
        print(f"📊 Excel file: {excel_path}")
        print(f"📊 Comparison items: {len(comparison_data)}")
        print(f"{'='*60}\n")
        
        # Store results
        from utils.progress import progress_store, progress_lock
        
        with progress_lock:
            if session_id not in progress_store:
                progress_store[session_id] = {}
            
            progress_store[session_id].update({
                "excel_path": excel_path,
                "preview_data": comparison_data,
                "filename": f"comparison_{file1_name.replace('.xlsx', '')}_{file2_name.replace('.xlsx', '')}.xlsx",
                "status": "completed",
                "progress": 100,
                "message": "✅ Comparison complete!"
            })
            
            print(f"✅ Stored in progress_store")

    except Exception as e:
        import traceback
        error_msg = str(e)
        print(f"\n{'='*60}")
        print(f"❌ ERROR: {error_msg}")
        traceback.print_exc()
        print(f"{'='*60}\n")
        
        update_progress(session_id, 0, f"❌ Error: {error_msg}")
        
        from utils.progress import progress_store, progress_lock
        with progress_lock:
            if session_id in progress_store:
                progress_store[session_id]["status"] = "failed"
                progress_store[session_id]["error"] = error_msg

    finally:
        # Cleanup uploaded files
        for path in [file1_path, file2_path]:
            if path and os.path.exists(path):
                os.remove(path)
        print("🧹 Temporary files cleaned up\n")


def parse_comparison_json(response: str) -> List[Dict]:
    """Parse and validate comparison JSON from LLM"""
    import re
    
    print(f"   📥 Parsing LLM response...")
    
    # Clean response
    cleaned = response.strip()
    cleaned = re.sub(r'^```json\s*', '', cleaned, flags=re.MULTILINE)
    cleaned = re.sub(r'^```\s*', '', cleaned, flags=re.MULTILINE)
    cleaned = re.sub(r'\s*```$', '', cleaned, flags=re.MULTILINE)
    cleaned = cleaned.strip()
    
    # Extract JSON array
    array_match = re.search(r'\[.*\]', cleaned, re.DOTALL)
    if array_match:
        cleaned = array_match.group(0)
    
    try:
        data = json.loads(cleaned)
        
        if isinstance(data, dict):
            data = [data]
        elif not isinstance(data, list):
            print(f"   ⚠️ Unexpected data type: {type(data)}")
            return []
        
        # Validate structure
        validated = []
        for item in data:
            if isinstance(item, dict):
                validated.append({
                    "category": str(item.get("category", "Comparison")).strip(),
                    "section": str(item.get("section", "")).strip(),
                    "guideline1_value": str(item.get("guideline1_value", "")).strip(),
                    "guideline2_value": str(item.get("guideline2_value", "")).strip(),
                    "difference": str(item.get("difference", "")).strip(),
                })
        
        print(f"   ✅ Parsed {len(validated)} comparison items")
        return validated
        
    except json.JSONDecodeError as e:
        print(f"   ❌ JSON parse error: {e}")
        print(f"   Content: {cleaned[:500]}...")
        return []


def create_comparison_excel(data: List[Dict], output_path: str, file1_name: str, file2_name: str):
    """Create formatted Excel file with comparison results"""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Comparison Results"
    
    # Styling
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    
    added_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    removed_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    modified_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
    
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Headers
    headers = ["Category", "Section", file1_name, file2_name, "Difference"]
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = border
    
    # Data rows
    for row_num, item in enumerate(data, 2):
        ws.cell(row=row_num, column=1).value = item.get("category", "")
        ws.cell(row=row_num, column=2).value = item.get("section", "")
        ws.cell(row=row_num, column=3).value = item.get("guideline1_value", "")
        ws.cell(row=row_num, column=4).value = item.get("guideline2_value", "")
        ws.cell(row=row_num, column=5).value = item.get("difference", "")
        
        # Apply borders and wrapping
        for col_num in range(1, 6):
            cell = ws.cell(row=row_num, column=col_num)
            cell.border = border
            cell.alignment = Alignment(wrap_text=True, vertical='top')
        
        # Color code by category
        category = item.get("category", "").lower()
        category_cell = ws.cell(row=row_num, column=1)
        
        if "add" in category or "new" in category:
            category_cell.fill = added_fill
        elif "remove" in category or "delete" in category:
            category_cell.fill = removed_fill
        elif "modif" in category or "change" in category:
            category_cell.fill = modified_fill
    
    # Column widths
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 30
    ws.column_dimensions['C'].width = 40
    ws.column_dimensions['D'].width = 40
    ws.column_dimensions['E'].width = 50
    
    # Freeze header
    ws.freeze_panes = 'A2'
    
    wb.save(output_path)
    print(f"✅ Comparison Excel created: {output_path}")